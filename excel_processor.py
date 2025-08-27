import openpyxl
from datetime import datetime

from constant import COUNTRY_COLUMN, DATE_COLUMN, INT_COLUMN, REQUIREMENT_COLUMNS


def excel_process(filter_column: str, filter_value: str, output_path: str, input_path: str) -> None:
    if not input_path or not output_path:
        raise ValueError("Нет входного или выходного файла!")

    if filter_column in INT_COLUMN:
        try:
            filter_value = int(filter_value)  # type: ignore
        except ValueError:
            raise ValueError(f"Ожидается число, но получено '{filter_value}'")

    if filter_column in DATE_COLUMN:
        try:
            checking = datetime.strptime(filter_value, "%d.%m.%Y").date()  # type: ignore
        except ValueError:
            raise ValueError("Ожидается дата в правильном формате!")

    if filter_column not in INT_COLUMN + DATE_COLUMN + COUNTRY_COLUMN:
        filter_value = filter_value.capitalize()
    elif filter_column in COUNTRY_COLUMN:
        filter_value = filter_value.upper()

    required_columns = REQUIREMENT_COLUMNS
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    headers = {}
    for col_idx, cell in enumerate(ws[9], 1):
        if cell.value:
            headers[cell.value.strip().lower()] = col_idx
    filter_column = headers[filter_column.lower()]

    cell_range = ws[
        ws.cell(row=11, column=1)
        .coordinate: ws.cell(row=ws.max_row, column=ws.max_column)
        .coordinate
    ]

    filter_data = []
    for row in cell_range:
        cell_value = row[filter_column - 1].value
        if cell_value == filter_value:
            filter_data.append([x.value for x in row])

    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.append(required_columns)
    important_columns = [v - 1 for k, v in headers.items() if k in required_columns]
    row_number = 2

    for row in filter_data:
        selected_columns = [row[i] for i in important_columns]
        for col_number, cell_value in enumerate(selected_columns, start=1):
            new_ws.cell(row=row_number, column=col_number, value=cell_value)
        row_number += 1

    new_wb.save(output_path)
    new_wb.close()
